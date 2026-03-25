import logging
import os
from datetime import datetime
from typing import Any, Dict, List

from app.services.bom_rules import canonicalize_row
from app.services.bom_schema import build_size_payload, ensure_list
from app.services.catia_bridge import catia_bridge
from app.services.tree_extractor import tree_extractor

logger = logging.getLogger(__name__)

OUTPUT_DIR = r"H:\CADMation\BOM_Outputs"


class BOMService:
    def _log_op(self, msg: str):
        """Logs an operation with timestamp to a dedicated file."""
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        log_file = os.path.join(OUTPUT_DIR, "operations.log")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {msg}\n"
        try:
            with open(log_file, "a", encoding="utf-8") as file_obj:
                file_obj.write(line)
        except Exception:
            pass
        logger.info(f"BOM_OP: {msg}")

    def _is_bom_candidate(self, node: Dict[str, Any]) -> bool:
        origin_type = node.get("originType")
        if origin_type == "assembly_container":
            return False
        return node.get("type") in ("Part", "Component") or origin_type in {"leaf_part", "std_leaf", "unknown_leaf"}

    def _normalize_sheet_category(self, row: Dict[str, Any]) -> str:
        category = row.get("exportBucket") or row.get("sheetCategory") or "Steel"
        if category in {"STD-MISUMI", "STD-OTHER"}:
            return "STD"
        if category in {"Steel", "MS", "Casting", "STD"}:
            return category
        return "Steel"

    def _extract_document_metadata(self, fallback_name: str = "BOM") -> Dict[str, str]:
        metadata = {
            "title": "CADMation Production BOM",
            "date": datetime.now().strftime("%d/%m/%Y"),
            "woNo": fallback_name,
            "customer": "-",
            "toolType": "-",
            "toolSize": "-",
            "projectName": fallback_name,
        }
        try:
            caa = catia_bridge.get_application()
            if not caa:
                return metadata
            doc = caa.ActiveDocument
            doc_name = getattr(doc, "Name", fallback_name).split(".")[0]
            metadata["projectName"] = doc_name or fallback_name
            metadata["woNo"] = doc_name or fallback_name
            product = getattr(doc, "Product", None)
            if product is not None:
                metadata["projectName"] = getattr(product, "PartNumber", "") or doc_name or fallback_name
                try:
                    params = getattr(product, "Parameters", None)
                    if params:
                        for param in params:
                            pname = getattr(param, "Name", "").upper()
                            value = getattr(param, "ValueAsString", lambda: "")()
                            if not value:
                                continue
                            if "CUSTOMER" in pname:
                                metadata["customer"] = value
                            elif "TOOL" in pname and "TYPE" in pname:
                                metadata["toolType"] = value
                            elif "TOOL" in pname and "SIZE" in pname:
                                metadata["toolSize"] = value
                            elif "WO" in pname:
                                metadata["woNo"] = value
                except Exception:
                    pass
        except Exception:
            pass
        return metadata

    def _build_row_from_node(self, node: Dict[str, Any], index: int) -> Dict[str, Any]:
        props = node.get("properties") or {}
        part_number = node.get("partNumber", node.get("name", "").split(".")[0]).strip()
        return canonicalize_row(
            {
                "id": node.get("id", index),
                "name": node.get("name", part_number),
                "partNumber": part_number,
                "instanceName": node.get("name", part_number),
                "qty": node.get("qty", 1),
                "instances": node.get("instances") or [node.get("name", part_number)],
                "selected": node.get("selected", True),
                "material": props.get("material", node.get("material", "")),
                "size": props.get("stock_size", node.get("size", "Not Measurable")),
                "millingSize": node.get("millingSize", props.get("stock_size", node.get("size", "Not Measurable"))),
                "rmSize": node.get("rmSize"),
                "heatTreatment": props.get("heat_treatment", node.get("heatTreatment", "NONE")),
                "methodUsed": node.get("methodUsed", props.get("method_used", "UNKNOWN")),
                "isStd": node.get("isStd", False),
                "manufacturer": node.get("manufacturer", ""),
                "description": node.get("description", ""),
                "remark": node.get("remark", ""),
                "sheetCategory": node.get("sheetCategory", ""),
                "catalogCode": node.get("catalogCode", ""),
                "machiningStock": node.get("machiningStock", 0),
                "roundingMm": node.get("roundingMm", 0),
                "validationFlags": node.get("validationFlags", []),
                "measurementConfidence": node.get("measurementConfidence", ""),
                "originType": node.get("originType", ""),
                "parentAssembly": node.get("parentAssembly", ""),
                "referenceKey": node.get("referenceKey", ""),
                "sourceDocPath": node.get("sourceDocPath", ""),
                "sourceDocumentName": node.get("sourceDocumentName", ""),
            }
        )

    def _collect_fast_items(self, node: Dict[str, Any], out: Dict[str, Dict[str, Any]]):
        name = node.get("name", "")
        node_type = node.get("type")
        part_number = node.get("partNumber", name.split(".")[0]).strip()
        if self._is_bom_candidate(node):
            if part_number in out:
                out[part_number]["qty"] += 1
                if name not in out[part_number]["instances"]:
                    out[part_number]["instances"].append(name)
            else:
                out[part_number] = {
                    "id": part_number,
                    "name": part_number,
                    "partNumber": part_number,
                    "instanceName": name,
                    "type": node_type,
                    "qty": 1,
                    "selected": True,
                    "instances": [name],
                    "originType": node.get("originType", ""),
                    "parentAssembly": node.get("parentAssembly", ""),
                    "referenceKey": node.get("referenceKey", ""),
                    "sourceDocPath": node.get("sourceDocPath", ""),
                }
        for child in node.get("children", []):
            self._collect_fast_items(child, out)

    def _collect_items(self, node: Dict[str, Any], out: List[Dict[str, Any]]):
        if self._is_bom_candidate(node):
            out.append(self._build_row_from_node(node, len(out) + 1))
        for child in node.get("children", []):
            self._collect_items(child, out)

    def _rollup_key(self, row: Dict[str, Any]) -> tuple:
        return (
            row.get("classification"),
            row.get("sheetCategory"),
            row.get("catalogCode"),
            row.get("partNumber"),
            row.get("description"),
            row.get("material"),
            row.get("manufacturer"),
            row.get("millingSize"),
            row.get("rmSize"),
            row.get("remark"),
            row.get("heatTreatment"),
            row.get("exportBucket"),
            row.get("reviewStatus"),
            row.get("discrepancyType"),
            row.get("originType"),
        )

    def _rollup_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        grouped: Dict[tuple, Dict[str, Any]] = {}
        for row in rows:
            canonical = canonicalize_row(row)
            canonical["sheetCategory"] = self._normalize_sheet_category(canonical)
            canonical["exportBucket"] = canonical["sheetCategory"]
            key = self._rollup_key(canonical)
            if key not in grouped:
                grouped[key] = {**canonical}
                grouped[key]["instances"] = ensure_list(canonical.get("instances")) or [canonical.get("instanceName")]
                continue
            target = grouped[key]
            target["qty"] += canonical.get("qty", 1)
            target["instances"] = sorted(set(target.get("instances", []) + ensure_list(canonical.get("instances"))))
            target["validationFlags"] = sorted(set(target.get("validationFlags", []) + canonical.get("validationFlags", [])))
            for weight_field in ("rmWeightKg", "finishWeightKg"):
                current = target.get(weight_field) or 0
                incoming = canonical.get(weight_field) or 0
                target[weight_field] = round(current + incoming, 3)
        return sorted(
            grouped.values(),
            key=lambda row: (
                row.get("sheetCategory", ""),
                row.get("parentAssembly", ""),
                row.get("description", ""),
                row.get("partNumber", ""),
            ),
        )

    def build_measured_row(self, base_item: Dict[str, Any], bbox: Dict[str, Any], effective_method: str) -> Dict[str, Any]:
        row = {
            **base_item,
            "partNumber": base_item.get("partNumber") or base_item.get("id") or base_item.get("name"),
            "size": bbox.get("stock_size", "Not Measurable"),
            "millingSize": bbox.get("stock_size", "Not Measurable"),
            "methodUsed": bbox.get("method_used", effective_method),
            "material": base_item.get("material", "STEEL"),
            "qty": base_item.get("qty", 1),
            "instances": base_item.get("instances") or [base_item.get("instanceName")],
            "originType": base_item.get("originType", "leaf_part"),
            "parentAssembly": base_item.get("parentAssembly", ""),
            "referenceKey": base_item.get("referenceKey", ""),
            "sourceDocPath": base_item.get("sourceDocPath", ""),
            "rawDims": bbox.get("rawDims", []),
            "orderedDims": bbox.get("orderedDims", []),
            "stockForm": bbox.get("stockForm", ""),
            "measurementConfidence": bbox.get("measurement_confidence", bbox.get("measurementConfidence", "")),
            "measurementBodyName": base_item.get("measurementBodyName", "")
            or base_item.get("roughStockBodyName", ""),
        }
        return canonicalize_row(row)

    def build_retry_candidate(self, item: Dict[str, Any]) -> Dict[str, Any]:
        return canonicalize_row(
            {
                "id": item.get("id"),
                "name": item.get("name"),
                "partNumber": item.get("partNumber") or item.get("id"),
                "instanceName": item.get("instanceName"),
                "instances": item.get("instances") or [item.get("instanceName")],
                "qty": item.get("qty", 1),
                "selected": True,
                "material": item.get("material", "STEEL"),
                "size": item.get("size", "Not Measurable"),
                "methodUsed": item.get("methodUsed", "ROUGH_STOCK"),
                "isStd": item.get("isStd", False),
                "manufacturer": item.get("manufacturer", ""),
                "description": item.get("description", ""),
                "remark": item.get("remark", ""),
                "sheetCategory": item.get("sheetCategory", ""),
                "originType": item.get("originType", ""),
                "parentAssembly": item.get("parentAssembly", ""),
                "referenceKey": item.get("referenceKey", ""),
                "sourceDocPath": item.get("sourceDocPath", ""),
            }
        )

    def get_bom_items(self) -> List[Dict[str, Any]]:
        self._log_op("Starting full BOM scan (properties included)...")
        tree = tree_extractor.get_full_tree(include_props=True, check_visibility=True)
        if not tree or "error" in tree:
            return []
        items: List[Dict[str, Any]] = []
        self._collect_items(tree, items)
        return self._rollup_rows(items)

    def get_bom_fast_list(self) -> List[Dict[str, Any]]:
        self._log_op("Starting fast BOM scan (selectable list)...")
        tree = tree_extractor.get_full_tree(include_props=False, check_visibility=True)
        if not tree or "error" in tree:
            return []
        grouped_items: Dict[str, Dict[str, Any]] = {}
        self._collect_fast_items(tree, grouped_items)
        rows = [canonicalize_row(item) for item in grouped_items.values()]
        for row in rows:
            if not row.get("isStd"):
                row["material"] = ""
        return rows

    def generate_excel_bom(self, edited_items: List[Dict[str, Any]] | None = None, check_visibility: bool = False) -> str | None:
        rows = edited_items if edited_items is not None else self.get_bom_items()
        if not rows:
            logger.warning("BOMService: No rows available for Excel generation.")
            return None
        rows = [canonicalize_row(row) for row in rows if row.get("keepInExport", row.get("selected", True))]
        if not rows:
            logger.warning("BOMService: All rows were filtered out before export.")
            return None
        project_name = rows[0].get("projectName", "BOM") if rows else "BOM"
        metadata = self._extract_document_metadata(project_name)
        return self._write_excel(self._rollup_rows(rows), metadata)

    def save_excel_bom(self, items: List[Dict[str, Any]]) -> str | None:
        return self.generate_excel_bom(edited_items=items)

    def _build_summary_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        summary: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            key = self._normalize_sheet_category(row)
            if key not in summary:
                summary[key] = {"category": key, "qty": 0, "rmWeightKg": 0.0, "finishWeightKg": 0.0}
            summary[key]["qty"] += row.get("qty", 0)
            summary[key]["rmWeightKg"] += row.get("rmWeightKg") or 0.0
            summary[key]["finishWeightKg"] += row.get("finishWeightKg") or 0.0
        return sorted(summary.values(), key=lambda item: item["category"])

    def _apply_styles(self, ws, title_fill, header_fill, thin_border, center_align, left_align, title_font, header_font, data_font):
        for cell in ws[1]:
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = center_align
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            is_section_row = row[0].row > 10 and row[0].value and all(cell.value in (None, "") for cell in row[1:])
            for cell in row:
                if cell.row in (9, 10):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                elif is_section_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_align if cell.column == 1 else center_align
                elif cell.row > 10:
                    cell.font = data_font
                    cell.alignment = (
                        center_align if cell.column in {1, 6, 7, 8, 9, 10, 11, 12, 13} else left_align
                    )
                cell.border = thin_border

    def _write_summary_sheet(self, wb, metadata: Dict[str, str], summary_rows: List[Dict[str, Any]], all_rows: List[Dict[str, Any]]):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        ws = wb.create_sheet(title="Sheet1")
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A1:N1")
        ws["A1"] = metadata["title"]
        ws["A3"] = "BOM"
        ws["B3"] = metadata["date"]
        ws["A4"] = "WO No :"
        ws["B4"] = metadata["woNo"]
        ws["A5"] = "Customer :"
        ws["B5"] = metadata["customer"]
        ws["A6"] = "Tool Type :"
        ws["B6"] = metadata["toolType"]
        ws["A7"] = "Tool Size :"
        ws["B7"] = metadata["toolSize"]
        ws["A8"] = "Summary"

        headers = ["Sr. No.", "Category", "Qty", "RM Weight (kg)", "Finish Weight (kg)"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=9, column=idx, value=header)

        for row_idx, row in enumerate(summary_rows, start=10):
            ws.cell(row=row_idx, column=1, value=row_idx - 9)
            ws.cell(row=row_idx, column=2, value=row["category"])
            ws.cell(row=row_idx, column=3, value=row["qty"])
            ws.cell(row=row_idx, column=4, value=round(row["rmWeightKg"], 3))
            ws.cell(row=row_idx, column=5, value=round(row["finishWeightKg"], 3))

        ws["H3"] = f"BOM Summary - {metadata['projectName']}"
        ws["H4"] = "Total Rows"
        ws["I4"] = len(all_rows)
        ws["H5"] = "STD Rows"
        ws["I5"] = sum(1 for row in all_rows if row.get("isStd"))
        ws["H6"] = "MFG Rows"
        ws["I6"] = sum(1 for row in all_rows if not row.get("isStd"))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                if cell.row == 1:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = center
                elif cell.row in (8, 9):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                else:
                    cell.font = data_font
                    cell.alignment = left if cell.column in {1, 2, 8} else center
                cell.border = thin_border

        for col_letter, width in {"A": 16, "B": 24, "C": 12, "D": 18, "E": 18, "H": 22, "I": 14}.items():
            ws.column_dimensions[col_letter].width = width

    def _write_mfg_sheet(self, wb, metadata: Dict[str, str], sheet_name: str, rows: List[Dict[str, Any]]):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        ws = wb.create_sheet(title=sheet_name)
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.merge_cells("A1:M1")
        ws["A1"] = metadata["title"]
        ws["A3"] = "BOM"
        ws["B3"] = metadata["date"]
        ws["A4"] = "WO No :"
        ws["B4"] = metadata["woNo"]
        ws["A5"] = "Customer :"
        ws["B5"] = metadata["customer"]
        ws["A6"] = "Tool Type :"
        ws["B6"] = metadata["toolType"]
        ws["A7"] = "Tool Size :"
        ws["B7"] = metadata["toolSize"]
        ws["A8"] = sheet_name

        merges = ["A9:A10", "B9:B10", "C9:C10", "D9:D10", "E9:E10", "F9:H9", "I9:I10", "J9:L9", "M9:M10"]
        for merge in merges:
            ws.merge_cells(merge)
        ws["A9"] = "Sr. No."
        ws["B9"] = "Description"
        ws["C9"] = "Part Number"
        ws["D9"] = "RM"
        ws["E9"] = "Remark"
        ws["F9"] = "Milling Size"
        ws["I9"] = "Qty"
        ws["J9"] = "RM Size"
        ws["M9"] = "RM Weight"
        ws["F10"] = "L"
        ws["G10"] = "W"
        ws["H10"] = "H"
        ws["J10"] = "L"
        ws["K10"] = "W"
        ws["L10"] = "H"

        current_row = 11
        current_section = None
        serial = 1
        for row in rows:
            next_section = row.get("parentAssembly") or sheet_name
            if next_section != current_section:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
                section_cell = ws.cell(row=current_row, column=1, value=next_section)
                section_cell.font = header_font
                section_cell.fill = header_fill
                section_cell.alignment = left
                current_row += 1
                current_section = next_section

            ws.cell(row=current_row, column=1, value=99 + serial)
            ws.cell(row=current_row, column=2, value=row.get("description"))
            ws.cell(row=current_row, column=3, value=row.get("partNumber"))
            ws.cell(row=current_row, column=4, value=row.get("material"))
            ws.cell(row=current_row, column=5, value=row.get("remark"))
            ws.cell(row=current_row, column=6, value=row.get("L"))
            ws.cell(row=current_row, column=7, value=row.get("W"))
            ws.cell(row=current_row, column=8, value=row.get("H"))
            ws.cell(row=current_row, column=9, value=row.get("qty"))
            ws.cell(row=current_row, column=10, value=row.get("rmL"))
            ws.cell(row=current_row, column=11, value=row.get("rmW"))
            ws.cell(row=current_row, column=12, value=row.get("rmH"))
            ws.cell(row=current_row, column=13, value=row.get("rmWeightKg"))
            if row.get("_MERGE_LW"):
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
            if row.get("rmMergeLW"):
                ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row, end_column=11)
            current_row += 1
            serial += 1

        self._apply_styles(ws, title_fill, header_fill, thin_border, center, left, title_font, header_font, data_font)
        widths = {"A": 10, "B": 28, "C": 22, "D": 12, "E": 18, "F": 11, "G": 11, "H": 11, "I": 8, "J": 11, "K": 11, "L": 11, "M": 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A11"

    def _write_std_sheet(self, wb, metadata: Dict[str, str], sheet_name: str, rows: List[Dict[str, Any]]):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        ws = wb.create_sheet(title=sheet_name)
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.merge_cells("A1:F1")
        ws["A1"] = metadata["title"]
        ws["A3"] = "BOM"
        ws["B3"] = metadata["date"]
        ws["A4"] = "WO No :"
        ws["B4"] = metadata["woNo"]
        ws["A5"] = "Customer :"
        ws["B5"] = metadata["customer"]
        ws["A6"] = "Tool Type :"
        ws["B6"] = metadata["toolType"]
        ws["A7"] = "Tool Size :"
        ws["B7"] = metadata["toolSize"]
        ws["A8"] = sheet_name

        headers = ["Sr. No.", "Description", "Type", "Manufacturer", "Catalog Code", "Qty"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=9, column=idx, value=header)

        for row_idx, row in enumerate(rows, start=10):
            ws.cell(row=row_idx, column=1, value=99 + row_idx - 9)
            ws.cell(row=row_idx, column=2, value=row.get("description"))
            ws.cell(row=row_idx, column=3, value="STD")
            ws.cell(row=row_idx, column=4, value=row.get("manufacturer"))
            ws.cell(row=row_idx, column=5, value=row.get("catalogCode") or row.get("partNumber"))
            ws.cell(row=row_idx, column=6, value=row.get("qty"))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                if cell.row == 1:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = center
                elif cell.row in (8, 9):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                else:
                    cell.font = data_font
                    cell.alignment = center if cell.column in {1, 3, 6} else left
                cell.border = thin_border

        for col, width in {"A": 10, "B": 26, "C": 10, "D": 16, "E": 20, "F": 8}.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A10"

    def _write_excel(self, rows: List[Dict[str, Any]], metadata: Dict[str, str]) -> str | None:
        from openpyxl import Workbook

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        name_base = metadata.get("projectName", "BOM").split(".")[0]
        filename = f"{name_base}_BOM_{timestamp}.xlsx"
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        file_path = os.path.join(OUTPUT_DIR, filename)

        try:
            wb = Workbook()
            wb.remove(wb.active)

            self._write_summary_sheet(wb, metadata, self._build_summary_rows(rows), rows)
            categories = {}
            for row in rows:
                category = self._normalize_sheet_category(row)
                categories.setdefault(category, []).append({**row, "sheetCategory": category, "exportBucket": category})

            for sheet_name in ["Steel", "MS", "Casting", "STD"]:
                if sheet_name not in categories:
                    continue
                if sheet_name == "STD":
                    self._write_std_sheet(wb, metadata, sheet_name, categories[sheet_name])
                else:
                    self._write_mfg_sheet(wb, metadata, sheet_name, categories[sheet_name])

            wb.save(file_path)
            self._log_op(f"BOM exported successfully to: {file_path}")
            logger.info(f"BOMService: Production BOM written to {file_path}")
            return file_path
        except Exception as exc:
            logger.error(f"BOMService: Failed to write Excel: {exc}")
            return None


bom_service = BOMService()
