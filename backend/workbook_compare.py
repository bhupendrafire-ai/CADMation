import argparse
import json
import re
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def normalize_text(value) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def extract_numbers(*values) -> List[float]:
    numbers: List[float] = []
    for value in values:
        if value in (None, ""):
            continue
        numbers.extend(float(match) for match in re.findall(r"-?\d+(?:\.\d+)?", str(value)))
    return numbers


def size_signature(values: List[float]) -> tuple:
    return tuple(sorted([round(abs(value), 1) for value in values if value is not None], reverse=True))


def inspect_workbook(path: str) -> Dict[str, object]:
    workbook = load_workbook(path)
    report: Dict[str, object] = {"path": str(path), "sheet_names": workbook.sheetnames, "sheets": {}}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        report["sheets"][sheet_name] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": sorted(str(rng) for rng in ws.merged_cells.ranges),
            "row_9": [ws.cell(row=9, column=col).value for col in range(1, ws.max_column + 1)],
            "row_10": [ws.cell(row=10, column=col).value for col in range(1, ws.max_column + 1)],
        }
    return report


def validate_workbook_structure(path: str) -> List[str]:
    report = inspect_workbook(path)
    errors: List[str] = []
    sheet_names = report["sheet_names"]
    if "Sheet1" not in sheet_names:
        errors.append("Missing summary sheet 'Sheet1'.")
    if not any(name in sheet_names for name in ("Steel", "MS", "Casting", "STD")):
        errors.append("Missing category sheets.")

    sheets = report["sheets"]
    for sheet_name in ("Steel", "MS", "Casting"):
        if sheet_name not in sheets:
            continue
        merged_ranges = set(sheets[sheet_name]["merged_ranges"])
        if "A9:A10" not in merged_ranges or "F9:H9" not in merged_ranges or "J9:L9" not in merged_ranges:
            errors.append(f"{sheet_name}: expected grouped headers are missing.")

    for sheet_name in ("STD",):
        if sheet_name not in sheets:
            continue
        row_9 = sheets[sheet_name]["row_9"]
        if "Catalog Code" not in row_9:
            errors.append(f"{sheet_name}: missing standard-part headers.")

    return errors


def parse_generated_rows(path: str) -> List[Dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    rows: List[Dict[str, object]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in {"Steel", "MS", "Casting", "STD"}:
            continue
        ws = workbook[sheet_name]
        start_row = 10 if sheet_name.startswith("STD") else 11
        for row_idx in range(start_row, ws.max_row + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if not any(value not in (None, "") for value in values):
                continue
            if sheet_name == "STD":
                if not values[0] or str(values[0]).strip() == "Sr. No.":
                    continue
                rows.append({
                    "sheet": sheet_name,
                    "description": values[1],
                    "catalog": values[4],
                    "qty": values[5],
                    "milling_dims": [],
                    "rm_dims": [],
                })
                continue
            if values[0] and all(value in (None, "") for value in values[1:]):
                continue
            rows.append({
                "sheet": sheet_name,
                "description": values[1],
                "part_number": values[2],
                "material": values[3],
                "remark": values[4],
                "milling_dims": extract_numbers(values[5], values[6], values[7]),
                "qty": values[8],
                "rm_dims": extract_numbers(values[9], values[10], values[11]),
            })
    return rows


def parse_engineer_rows(path: str) -> List[Dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    rows: List[Dict[str, object]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in {"Steel", "MS", "Casting", "STD", "STD-MISUMI", "STD-OTHER"}:
            continue
        ws = workbook[sheet_name]
        normalized_sheet = "STD" if sheet_name.startswith("STD") else sheet_name
        start_row = 10 if normalized_sheet == "STD" else 11
        for row_idx in range(start_row, ws.max_row + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if not any(value not in (None, "") for value in values):
                continue
            first = str(values[0] or "").strip()
            if normalized_sheet == "STD":
                if not first.startswith("PTE"):
                    continue
                rows.append({
                    "sheet": normalized_sheet,
                    "description": values[1],
                    "catalog": values[3],
                    "qty": values[4],
                    "milling_dims": [],
                    "rm_dims": [],
                })
                continue
            if not first.startswith("PTE"):
                continue
            milling = extract_numbers(values[4], values[5], values[6])
            rm = extract_numbers(values[7], values[8], values[9], values[10])
            if not rm:
                rm = extract_numbers(values[7], values[8], values[9])
            rows.append({
                "sheet": normalized_sheet,
                "description": values[1],
                "material": values[2],
                "remark": values[3],
                "milling_dims": milling,
                "rm_dims": rm,
                "qty": values[11] if len(values) > 11 else values[4],
                "catalog": values[3] if normalized_sheet == "STD" else "",
            })
    return rows


def row_match_score(reference_row: Dict[str, object], candidate_row: Dict[str, object]) -> int:
    score = 0
    ref_desc = normalize_text(reference_row.get("description"))
    cand_desc = normalize_text(candidate_row.get("description")) + " " + normalize_text(candidate_row.get("part_number")) + " " + normalize_text(candidate_row.get("catalog"))
    ref_catalog = normalize_text(reference_row.get("catalog"))
    cand_catalog = normalize_text(candidate_row.get("catalog"))

    if ref_catalog and cand_catalog and ref_catalog == cand_catalog:
        score += 6
    if ref_desc and ref_desc in cand_desc:
        score += 4
    ref_sig = size_signature(reference_row.get("milling_dims", []))
    cand_sig = size_signature(candidate_row.get("milling_dims", []))
    if ref_sig and ref_sig == cand_sig:
        score += 5
    elif ref_sig and cand_sig and set(ref_sig) == set(cand_sig):
        score += 3
    if reference_row.get("sheet") == candidate_row.get("sheet"):
        score += 1
    return score


def semantic_compare_rows(reference_rows: List[Dict[str, object]], candidate_rows: List[Dict[str, object]]) -> Dict[str, object]:
    unmatched_candidates = candidate_rows.copy()
    exact_matches = []
    probable_matches = []
    missing_rows = []
    dimension_mismatches = []
    sheet_mismatches = []

    for reference_row in reference_rows:
        scored = sorted(
            ((row_match_score(reference_row, candidate_row), candidate_row) for candidate_row in unmatched_candidates),
            key=lambda item: item[0],
            reverse=True,
        )
        best_score, best_candidate = scored[0] if scored else (0, None)
        if best_score < 4 or best_candidate is None:
            missing_rows.append(reference_row)
            continue

        unmatched_candidates.remove(best_candidate)
        payload = {"reference": reference_row, "candidate": best_candidate, "score": best_score}
        ref_sig = size_signature(reference_row.get("milling_dims", []))
        cand_sig = size_signature(best_candidate.get("milling_dims", []))
        if reference_row.get("sheet") != best_candidate.get("sheet"):
            sheet_mismatches.append(payload)
        if ref_sig and cand_sig and ref_sig != cand_sig:
            dimension_mismatches.append(payload)
        if best_score >= 9 and ref_sig == cand_sig:
            exact_matches.append(payload)
        else:
            probable_matches.append(payload)

    return {
        "exact_matches": exact_matches,
        "probable_matches": probable_matches,
        "missing_rows": missing_rows,
        "extra_rows": unmatched_candidates,
        "dimension_mismatches": dimension_mismatches,
        "sheet_mismatches": sheet_mismatches,
    }


def compare_workbooks(reference_path: str, candidate_path: str) -> Dict[str, object]:
    reference = inspect_workbook(reference_path)
    candidate = inspect_workbook(candidate_path)
    shared_sheets = sorted(set(reference["sheet_names"]) & set(candidate["sheet_names"]))
    row_analysis = semantic_compare_rows(parse_engineer_rows(reference_path), parse_generated_rows(candidate_path))
    return {
        "reference": str(reference_path),
        "candidate": str(candidate_path),
        "shared_sheets": shared_sheets,
        "reference_only": sorted(set(reference["sheet_names"]) - set(candidate["sheet_names"])),
        "candidate_only": sorted(set(candidate["sheet_names"]) - set(reference["sheet_names"])),
        "row_analysis": {
            "exact_match_count": len(row_analysis["exact_matches"]),
            "probable_match_count": len(row_analysis["probable_matches"]),
            "missing_count": len(row_analysis["missing_rows"]),
            "extra_count": len(row_analysis["extra_rows"]),
            "dimension_mismatch_count": len(row_analysis["dimension_mismatches"]),
            "sheet_mismatch_count": len(row_analysis["sheet_mismatches"]),
            "missing_rows": row_analysis["missing_rows"][:10],
            "extra_rows": row_analysis["extra_rows"][:10],
            "dimension_mismatches": row_analysis["dimension_mismatches"][:10],
            "sheet_mismatches": row_analysis["sheet_mismatches"][:10],
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Inspect or compare generated BOM workbooks.")
    parser.add_argument("candidate", help="Generated workbook path")
    parser.add_argument("--reference", help="Engineer workbook path for sheet-level comparison")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of plain text")
    args = parser.parse_args()

    errors = validate_workbook_structure(args.candidate)
    payload: Dict[str, object] = {
        "candidate": str(Path(args.candidate)),
        "errors": errors,
        "structure": inspect_workbook(args.candidate),
    }
    if args.reference:
        payload["comparison"] = compare_workbooks(args.reference, args.candidate)

    if args.json:
        print(json.dumps(payload, indent=2))
        return

    print(f"Workbook: {args.candidate}")
    print("Status: PASS" if not errors else "Status: FAIL")
    for error in errors:
        print(f"- {error}")
    if args.reference:
        comparison = payload["comparison"]
        print(f"Shared sheets: {', '.join(comparison['shared_sheets']) or '(none)'}")
        if comparison["reference_only"]:
            print(f"Missing vs reference: {', '.join(comparison['reference_only'])}")
        if comparison["candidate_only"]:
            print(f"Extra vs reference: {', '.join(comparison['candidate_only'])}")
        row_analysis = comparison["row_analysis"]
        print(f"Exact row matches: {row_analysis['exact_match_count']}")
        print(f"Probable row matches: {row_analysis['probable_match_count']}")
        print(f"Missing rows: {row_analysis['missing_count']}")
        print(f"Extra rows: {row_analysis['extra_count']}")
        print(f"Dimension mismatches: {row_analysis['dimension_mismatch_count']}")


if __name__ == "__main__":
    main()
